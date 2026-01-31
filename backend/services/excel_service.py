"""
Excel (.xlsx) export for daily recruitment log, interview results, audit logs.
"""
from datetime import datetime
from io import BytesIO
from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from pymongo.database import Database

from database import CANDIDATES, INTERVIEWS, AUDIT_LOGS, USERS


HEADER_FILL = PatternFill(start_color="0066B3", end_color="0066B3", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def _apply_header(ws, row: int, headers: list):
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=col, value=h)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = HEADER_ALIGN
        c.border = BORDER


def daily_recruitment_log(
    db: Database,
    from_date: Optional[datetime] = None,
    to_date: Optional[datetime] = None,
) -> BytesIO:
    """Excel: Registration date, Interview date, Interviewer, Decision, Status."""
    q: dict = {}
    if from_date:
        q["created_at"] = {"$gte": from_date}
    if to_date:
        end = to_date.replace(hour=23, minute=59, second=59, microsecond=999999)
        q.setdefault("created_at", {})["$lte"] = end
    candidates = list(db[CANDIDATES].find(q).sort("created_at", -1))

    wb = Workbook()
    ws = wb.active
    ws.title = "Daily Recruitment Log"
    headers = [
        "Candidate ID",
        "Name",
        "Role Applied",
        "Registration Date",
        "Interview Date",
        "Interviewer",
        "Decision",
        "Status",
        "Eligibility",
    ]
    _apply_header(ws, 1, headers)

    row = 2
    for c in candidates:
        cand_oid = c["_id"]
        inv = db[INTERVIEWS].find_one({"candidate_oid": cand_oid}, sort=[("interview_date", -1)])
        interview_date = inv.get("interview_date") if inv else None
        interviewer_name = ""
        decision = ""
        if inv:
            u = db[USERS].find_one({"_id": inv["interviewer_id"]})
            interviewer_name = u.get("full_name", "") if u else ""
            decision = inv.get("decision", "")

        ca = c.get("created_at")
        ws.cell(row=row, column=1, value=c.get("candidate_id", ""))
        ws.cell(row=row, column=2, value=c.get("name", ""))
        ws.cell(row=row, column=3, value=c.get("role_applied") or "")
        ws.cell(row=row, column=4, value=ca.strftime("%Y-%m-%d %H:%M") if ca else "")
        ws.cell(row=row, column=5, value=interview_date.strftime("%Y-%m-%d %H:%M") if interview_date else "")
        ws.cell(row=row, column=6, value=interviewer_name)
        ws.cell(row=row, column=7, value=decision)
        ws.cell(row=row, column=8, value=c.get("status", ""))
        ws.cell(row=row, column=9, value=c.get("eligibility") or "")
        for col in range(1, 10):
            ws.cell(row=row, column=col).border = BORDER
        row += 1

    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 18

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def interview_results(
    db: Database,
    from_date: Optional[datetime] = None,
    to_date: Optional[datetime] = None,
    role: Optional[str] = None,
    decision: Optional[str] = None,
) -> BytesIO:
    """Excel: Interview results with filters."""
    q: dict = {}
    if from_date:
        q["interview_date"] = {"$gte": from_date}
    if to_date:
        end = to_date.replace(hour=23, minute=59, second=59, microsecond=999999)
        q.setdefault("interview_date", {})["$lte"] = end
    if decision:
        q["decision"] = decision
    if role:
        cand_ids = [d["_id"] for d in db[CANDIDATES].find({"role_applied": {"$regex": role, "$options": "i"}}, {"_id": 1})]
        q["candidate_oid"] = {"$in": cand_ids}
    interviews = list(db[INTERVIEWS].find(q).sort("interview_date", -1))

    wb = Workbook()
    ws = wb.active
    ws.title = "Interview Results"
    headers = [
        "Candidate ID",
        "Name",
        "Role Applied",
        "Interview Date",
        "Interviewer",
        "Decision",
        "Notes",
    ]
    _apply_header(ws, 1, headers)

    row = 2
    for i in interviews:
        c = db[CANDIDATES].find_one({"_id": i["candidate_oid"]}) or {}
        u = db[USERS].find_one({"_id": i["interviewer_id"]}) or {}
        idt = i.get("interview_date")
        ws.cell(row=row, column=1, value=c.get("candidate_id", ""))
        ws.cell(row=row, column=2, value=c.get("name", ""))
        ws.cell(row=row, column=3, value=c.get("role_applied") or "")
        ws.cell(row=row, column=4, value=idt.strftime("%Y-%m-%d %H:%M") if idt else "")
        ws.cell(row=row, column=5, value=u.get("full_name", ""))
        ws.cell(row=row, column=6, value=i.get("decision", ""))
        ws.cell(row=row, column=7, value=(i.get("notes") or "")[:500])
        for col in range(1, 8):
            ws.cell(row=row, column=col).border = BORDER
        row += 1

    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 20

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def audit_logs_report(
    db: Database,
    from_date: Optional[datetime] = None,
    to_date: Optional[datetime] = None,
) -> BytesIO:
    """Excel: Admin audit logs."""
    q: dict = {}
    if from_date:
        q["created_at"] = {"$gte": from_date}
    if to_date:
        end = to_date.replace(hour=23, minute=59, second=59, microsecond=999999)
        q.setdefault("created_at", {})["$lte"] = end
    logs = list(db[AUDIT_LOGS].find(q).sort("created_at", -1))

    wb = Workbook()
    ws = wb.active
    ws.title = "Audit Logs"
    headers = ["Timestamp", "User", "Action", "Resource Type", "Resource ID", "Details"]
    _apply_header(ws, 1, headers)

    row = 2
    for log in logs:
        user_email = "System"
        if log.get("user_id"):
            u = db[USERS].find_one({"_id": log["user_id"]})
            user_email = u.get("email", "System") if u else "System"
        ca = log.get("created_at")
        ws.cell(row=row, column=1, value=ca.strftime("%Y-%m-%d %H:%M:%S") if ca else "")
        ws.cell(row=row, column=2, value=user_email)
        ws.cell(row=row, column=3, value=log.get("action", ""))
        ws.cell(row=row, column=4, value=log.get("resource_type") or "")
        ws.cell(row=row, column=5, value=log.get("resource_id") or "")
        ws.cell(row=row, column=6, value=str(log.get("details"))[:500] if log.get("details") else "")
        for col in range(1, 7):
            ws.cell(row=row, column=col).border = BORDER
        row += 1

    for col in range(1, 7):
        ws.column_dimensions[get_column_letter(col)].width = 22

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
